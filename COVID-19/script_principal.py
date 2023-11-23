import requests
import matplotlib.pyplot as plt
import openpyxl
import os
import sys
import modulo_uno  # barras.py
import modulo_dos  # estadisticas.py

# Definir un diccionario para almacenar archivos guardados y sus rutas
archivos_guardados = {}

def obtener_datos_casos():
    url_casos = "https://disease.sh/v3/covid-19/all"
    response_casos = requests.get(url_casos)
    datos_casos = response_casos.json()
    return datos_casos

def obtener_datos_paises():
    url_paises = "https://disease.sh/v3/covid-19/countries"
    response_paises = requests.get(url_paises)
    datos_paises = response_paises.json()
    return datos_paises

def consultar_api(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f'Error al realizar la solicitud: {e}')
        return None

def obtener_ruta_guardado():
    return os.path.dirname(os.path.abspath(__file__))

# Otras definiciones de funciones...

def guardar_excel(data, filename):
    # Añadir la extensión .xlsx si no está presente
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    ruta = obtener_ruta_guardado()
    ruta_completa = os.path.join(ruta, filename)

    wb = openpyxl.Workbook()
    ws = wb.active

    if isinstance(data, dict):
        headers = list(data.keys())
        ws.append(headers)
        row = [str(data[header]) for header in headers]  # Convertir los valores a cadena de texto
        ws.append(row)
    else:
        print("Formato de datos no soportado para Excel.")
        return

    wb.save(ruta_completa)
    print("Archivo Excel guardado en:", ruta_completa)

    # Almacenar el archivo guardado y su ruta en el diccionario
    archivos_guardados[filename] = ruta_completa

def mostrar_archivos_guardados():
    if archivos_guardados:
        print("\nArchivos guardados:")
        for nombre_archivo in archivos_guardados.keys():
            print(nombre_archivo)  # Mostrar solo el nombre del archivo
    else:
        print("\nNo hay archivos guardados aún.")

def leer_excel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    return [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

def leer_archivo_guardado():
    if not archivos_guardados:
        print("No hay archivos guardados.")
        return

    print("\nArchivos guardados:")
    archivos = list(archivos_guardados.keys())
    for i, nombre_archivo in enumerate(archivos, start=1):
        print(f"{i}. {nombre_archivo}")

    seleccion = input("Seleccione un archivo por número (1-{0}): ".format(len(archivos)))
    try:
        seleccion = int(seleccion)
        if 1 <= seleccion <= len(archivos):
            archivo_seleccionado = archivos[seleccion - 1]
            extension = os.path.splitext(archivo_seleccionado)[1]
            if extension == '.xlsx':
                data = leer_excel(archivos_guardados[archivo_seleccionado])
                print(f"Datos leídos del archivo Excel '{archivo_seleccionado}':")
                for entry in data:
                    print(entry)
            else:
                print("Formato de archivo no compatible.")
        else:
            print("Selección no válida.")
    except ValueError:
        print("Entrada no válida. Ingrese un número válido.")

def main():
    data = None
    url1 = 'https://disease.sh/v3/covid-19/all'
    url2 = 'https://disease.sh/v3/covid-19/vaccine/coverage/countries?lastdays=1'

    while True:
        print("\nMenu Principal:")
        print("1. Consultar información de la API 1")
        print("2. Consultar información de la API 2")
        print("3. Guardar información en archivo Excel")
        print("4. Mostrar lista de archivos guardados")
        print("5. Leer archivo guardado")
        print("6. Crear Gráficos")
        print("7. Consultar datos estadísticos")  
        print("8. Salir")

        opcion = input("Seleccione una opción (1-8): ")

        if opcion == '1':
            data = consultar_api(url1)
            if data:
                print("Información de la API 1:")
                for key, value in data.items():
                    print(f"{key}: {value}")

        elif opcion == '2':
            data = consultar_api(url2)
            if data:
                formatted_data = {}  # Crear un diccionario vacío para almacenar los datos
                for entry in data:
                    country_name = entry['country']
                    timeline_data = entry['timeline']
                    if '11/22/23' in timeline_data:
                        valor = timeline_data['11/22/23']
                    else:
                        valor = "Fecha no encontrada"
                    formatted_data[country_name] = valor  # Agregar los datos al diccionario formateado
                filename = input("Ingrese el nombre del archivo Excel para guardar: ")
                guardar_excel(formatted_data, filename)
                print(f"Datos guardados en {filename}")

        elif opcion == '3':
            if data:
                filename = input("Ingrese el nombre del archivo Excel para guardar: ")
                guardar_excel(data, filename)
                print(f"Datos guardados en {filename}")

        elif opcion == '4':
            mostrar_archivos_guardados()

        elif opcion == '5':
            leer_archivo_guardado()
            
        elif opcion == '6':
            modulo_uno.main()
            
        elif opcion == '7':
            modulo_dos.main()

        elif opcion == '8':
            print("Saliendo del programa. ¡Hasta luego!")
            sys.exit()

        else:
            print("Opción no válida. Intente nuevamente.")

if __name__ == "__main__":
    main()
